"""Master Excel workbook read/write: one sheet per (abf, channel) session,
plus an aggregated "Summary" sheet recomputed from all other sheets."""

import os
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SUMMARY_SHEET_NAME = "Summary" #important in the LSV file

CHANNEL_SUFFIX = {"hippocampus": "_hpc", "thalamus": "_thal"}


@dataclass
class FileResult:
    abf_stem: str
    channel_name: str
    seizure_present: bool
    seizure_onset_s: Optional[float]
    num_gt_events: int
    num_algo_events: int
    tp: int
    fp: int
    fn: int
    sensitivity: Optional[float]
    tp_pairs: list  # list of (gt_time_s, algo_time_s)
    tn: int = 0  # true negatives (rejected events not detected by algo)
    fp_rejected: int = 0  # false positives from rejected events (algo detected something rejected)
    specificity: Optional[float] = None


def sheet_name_for(result: FileResult) -> str:
    suffix = CHANNEL_SUFFIX[result.channel_name]
    name = f"{result.abf_stem}{suffix}"
    # Excel sheet names are capped at 31 characters and cannot contain []:*?/\\
    for ch in "[]:*?/\\":
        name = name.replace(ch, "_")
    return name[:31]


def open_or_create_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        try:
            return load_workbook(path)
        except Exception as e:
            # File exists but is corrupted or not a valid xlsx file
            raise ValueError(
                f"Could not open '{os.path.basename(path)}'. "
                f"The file may be corrupted or not a valid Excel file. "
                f"Please select a different file or create a new one."
            ) from e

    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = SUMMARY_SHEET_NAME
    return wb


def write_file_sheet(wb: Workbook, result: FileResult) -> str:
    name = sheet_name_for(result)
    if name in wb.sheetnames:
        del wb[name]
    ws: Worksheet = wb.create_sheet(title=name)

    rows = [
        ("ABF File", result.abf_stem),
        ("Channel", result.channel_name),
        ("Seizure Present", "Yes" if result.seizure_present else "No"),
        ("Seizure Onset (s)", result.seizure_onset_s if result.seizure_onset_s is not None else "N/A"),
        ("Ground Truth Events (n)", result.num_gt_events),
        ("Algorithm Events, Pre-Seizure (n)", result.num_algo_events),
        ("TP (Validated & Algorithm Matched)", result.tp),
        ("FP (Algorithm Only)", result.fp),
        ("FN (Validated Only)", result.fn),
        ("Rejected Events (n)", result.fp_rejected),
        ("Sensitivity", result.sensitivity if result.sensitivity is not None else "N/A"),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)

    #Note this should not just be true positive events, it should be all events with an extra "feature" 
    #of whether or not it was real
    header_row = len(rows) + 3
    ws.cell(row=header_row - 1, column=1, value="True Positive Events")
    ws.cell(row=header_row, column=1, value="GT Time (s)")
    ws.cell(row=header_row, column=2, value="Matched Algo Time (s)")
    ws.cell(row=header_row, column=3, value="Features (TBD)")

    #Looks like row 17 is the first one with real data, start collecting from here for LSV and RFC

    for i, (gt_t, algo_t) in enumerate(result.tp_pairs, start=header_row + 1):
        ws.cell(row=i, column=1, value=gt_t)
        ws.cell(row=i, column=2, value=algo_t)

    for col, width in ((1, 26), (2, 26), (3, 20)):
        ws.column_dimensions[chr(64 + col)].width = width

    return name


def update_summary_sheet(wb: Workbook) -> None:
    if SUMMARY_SHEET_NAME in wb.sheetnames:
        del wb[SUMMARY_SHEET_NAME]
    ws = wb.create_sheet(title=SUMMARY_SHEET_NAME, index=0)

    recording_names = [n for n in wb.sheetnames if n != SUMMARY_SHEET_NAME]

    total_tp = total_fp = total_fn = total_fp_rejected = 0
    per_recording = []
    for name in recording_names:
        sh = wb[name]
        tp = sh.cell(row=7, column=2).value or 0
        fp = sh.cell(row=8, column=2).value or 0
        fn = sh.cell(row=9, column=2).value or 0
        fp_rejected = sh.cell(row=10, column=2).value or 0
        sens = sh.cell(row=11, column=2).value
        total_tp += tp
        total_fp += fp
        total_fn += fn
        total_fp_rejected += fp_rejected
        per_recording.append((name, tp, fp, fn, fp_rejected, sens))

    agg_sensitivity = (total_tp / (total_tp + total_fn)) if (total_tp + total_fn) > 0 else "N/A"

    meta_rows = [
        ("Number of Recordings", len(recording_names)),
        ("Last Updated", datetime.now().isoformat(timespec="seconds")),
        ("Aggregate TP", total_tp),
        ("Aggregate FP", total_fp),
        ("Aggregate FN", total_fn),
        ("Aggregate Rejected Events", total_fp_rejected),
        ("Aggregate Sensitivity", agg_sensitivity),
    ]
    for i, (label, value) in enumerate(meta_rows, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)

    header_row = len(meta_rows) + 3
    headers = ["Recording", "TP", "FP", "FN", "Rejected", "Sensitivity"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)

    for i, (name, tp, fp, fn, fp_rejected, sens) in enumerate(sorted(per_recording), start=header_row + 1):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=tp)
        ws.cell(row=i, column=3, value=fp)
        ws.cell(row=i, column=4, value=fn)
        ws.cell(row=i, column=5, value=fp_rejected)
        ws.cell(row=i, column=6, value=sens if sens is not None else "N/A")

    for col, width in ((1, 30), (2, 10), (3, 10), (4, 10), (5, 10), (6, 14), (7, 14)):
        ws.column_dimensions[chr(64 + col)].width = width


def save_result(path: str, result: FileResult) -> None:
    wb = open_or_create_workbook(path)
    write_file_sheet(wb, result)
    update_summary_sheet(wb)
    wb.save(path)
